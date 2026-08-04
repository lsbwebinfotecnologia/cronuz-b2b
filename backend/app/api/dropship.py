"""
API Dropship — Integração Hub-Erdos / Hórus ERP
Prefixo: /dropship
Versão API Hórus mínima: v.01.127

Fluxo completo:
1. Configurar token Erdos + customer vinculado + params fiscais em POST /dropship/config/{company_id}
2. Sincronizar pedidos: POST /dropship/orders/{company_id}/sync
   → Baixa pedidos do Hub, armazena documentos localmente (URLs expiram em 1h)
3. Enviar pedido ao Hórus: POST /dropship/orders/{company_id}/{order_id}/send-to-horus
   → Cria 2 pedidos no Hórus: remessa (6.923 c/ baixa estoque) + venda (6.118 s/ baixa estoque)
4. Confirmar despacho: POST /dropship/orders/{company_id}/{order_id}/confirm-dispatch
   → Envia tracking + chave NF-e 6.923 para o Hub-Erdos
"""
import os
import re
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Any, Dict

log = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.core.dependencies import get_current_user
from app.models.user import User, UserRole
from app.models.dropship import DropshipConfig, DropshipOrder
from app.models.dropship_price_table import DropshipPriceTable
from app.models.company import Company
from app.models.company_settings import CompanySettings
from app.models.customer import Customer
from app.integrators.erdos_client import ErdosClient, ErdosClientError
from app.integrators.horus_logistics import HorusLogisticsClient
from fastapi import UploadFile, File
from datetime import date
import io

router = APIRouter()


# ==============================================================================
# HELPERS
# ==============================================================================

def _require_master(current_user: User):
    if current_user.type not in [UserRole.MASTER]:
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores.")


def _require_seller_or_master(current_user: User, company_id: int):
    """Permite MASTER acessar qualquer company; SELLER apenas a própria."""
    if current_user.type == UserRole.MASTER:
        return
    if current_user.type == UserRole.SELLER and current_user.company_id == company_id:
        return
    raise HTTPException(status_code=403, detail="Acesso negado.")


def _get_config_or_404(db: Session, company_id: int) -> DropshipConfig:
    config = db.query(DropshipConfig).filter(
        DropshipConfig.company_id == company_id,
        DropshipConfig.provider == "ERDOS"
    ).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuração Dropship não encontrada para esta empresa.")
    return config


def _build_erdos_client(config: DropshipConfig) -> ErdosClient:
    if not config.enabled:
        raise HTTPException(status_code=400, detail="Integração Dropship está desativada para esta empresa.")
    if not config.api_token or not config.api_base_url:
        raise HTTPException(status_code=400, detail="Token ou URL base do Hub-Erdos não configurados.")
    return ErdosClient(base_url=config.api_base_url, api_key=config.api_token)


async def _download_and_store(client: ErdosClient, url: Optional[str], dest_path: str) -> Optional[str]:
    """
    Baixa um arquivo de URL assinada e salva localmente.
    Retorna o caminho relativo para armazenamento no banco.
    URLs do Hub-Erdos expiram em 1 hora — download deve ser imediato.
    """
    if not url:
        return None
    try:
        content = await client.download_file(url)
        Path(dest_path).parent.mkdir(parents=True, exist_ok=True)
        with open(dest_path, "wb") as f:
            f.write(content)
        return dest_path
    except ErdosClientError:
        return None


# ==============================================================================
# SCHEMAS
# ==============================================================================

class DropshipConfigCreate(BaseModel):
    provider: str = "ERDOS"
    enabled: bool = False
    api_token: Optional[str] = None
    api_base_url: Optional[str] = "https://wxcapqbtvgttooamglxx.supabase.co/functions/v1/api-fornecedor"
    horus_customer_id: Optional[int] = None
    horus_customer_cod_cli: Optional[str] = None
    # Parâmetros fiscais
    horus_fiscal_param_remessa_intra: Optional[str] = None
    horus_fiscal_param_remessa_inter: Optional[str] = None
    horus_fiscal_param_venda: Optional[str] = None
    # Parâmetros do cliente (InsCliente / InsPedidoVenda)
    horus_tipo_cliente: Optional[str] = None
    horus_resp_cliente: Optional[str] = None
    horus_cod_resp: Optional[str] = None
    horus_cod_endereco: Optional[str] = None
    # Parâmetros do pedido
    horus_cod_metodo: Optional[str] = None
    horus_cod_endereco_pedido: Optional[str] = None
    # Parâmetros exclusivos da Remessa B2C
    horus_cod_transp: Optional[str] = None
    horus_frete_emit_dest: Optional[str] = None
    horus_status_envio_erp: Optional[str] = None
    # Parâmetros financeiros dos pedidos Hórus
    vlr_taxa_frete: Optional[float] = 0.0
    perc_desconto_remessa: Optional[float] = 0.0
    # Sincronização de estoque
    stock_sync_interval_min: int = 30
    stock_sync_enabled: bool = False


class DropshipConfigResponse(BaseModel):
    id: int
    company_id: int
    provider: str
    enabled: bool
    api_token: Optional[str]
    api_base_url: Optional[str]
    horus_customer_id: Optional[int]
    horus_customer_name: Optional[str] = None
    horus_customer_document: Optional[str] = None
    horus_customer_id_guid: Optional[str] = None
    horus_customer_id_doc: Optional[str] = None
    horus_customer_cod_cli: Optional[str] = None
    # Parâmetros fiscais
    horus_fiscal_param_remessa: Optional[str] = None   # legado
    horus_fiscal_param_remessa_intra: Optional[str] = None
    horus_fiscal_param_remessa_inter: Optional[str] = None
    horus_fiscal_param_venda: Optional[str]
    # Parâmetros do cliente
    horus_tipo_cliente: Optional[str] = None
    horus_resp_cliente: Optional[str] = None
    horus_cod_resp: Optional[str] = None
    horus_cod_endereco: Optional[str] = None
    # Parâmetros do pedido
    horus_cod_metodo: Optional[str] = None
    horus_cod_endereco_pedido: Optional[str] = None
    # Parâmetros exclusivos da Remessa B2C
    horus_cod_transp: Optional[str] = None
    horus_frete_emit_dest: Optional[str] = None
    horus_status_envio_erp: Optional[str] = None
    # Parâmetros financeiros dos pedidos Hórus
    vlr_taxa_frete: Optional[float] = None
    perc_desconto_remessa: Optional[float] = None
    # Estoque
    stock_sync_interval_min: int
    stock_sync_enabled: bool
    stock_sync_last_run: Optional[datetime]
    created_at: Optional[datetime]
    updated_at: Optional[datetime]

    class Config:
        from_attributes = True


class DropshipOrderResponse(BaseModel):
    id: int
    company_id: int
    config_id: int
    external_order_id: str
    external_reference: Optional[str]
    channel: Optional[str]
    status: str
    released_at: Optional[datetime]
    customer_data: Optional[Any]
    items_data: Optional[Any]
    logistics_data: Optional[Any]
    fiscal_data: Optional[Any]
    horus_pedido_remessa: Optional[str]
    horus_pedido_venda: Optional[str]
    horus_cod_cli_final: Optional[str] = None   # COD_CLI do cliente final no Hórus
    tracking_code: Optional[str]
    nfe_remessa_key: Optional[str]
    label_path: Optional[str]
    danfe_path: Optional[str]
    xml_path: Optional[str]
    synced_at: Optional[datetime]
    sent_to_horus_at: Optional[datetime]
    dispatched_at: Optional[datetime]
    created_at: Optional[datetime]
    updated_at: Optional[datetime]
    # Campos de integração Erdos em tempo real
    erdos_status: Optional[str]
    erdos_checked_at: Optional[datetime]
    erdos_alert: Optional[bool]
    logs: Optional[Any]  # list[dict]
    conference: Optional[Any] = None

    class Config:
        from_attributes = True


class ConfirmDispatchRequest(BaseModel):
    tracking_code: Optional[str] = None
    nfe_remessa_key: Optional[str] = None


class CustomerSearchResult(BaseModel):
    id: int
    name: str
    document: str
    id_guid: Optional[str]
    id_doc: Optional[str]

    class Config:
        from_attributes = True


# ==============================================================================
# CONFIGURAÇÃO
# ==============================================================================

@router.get("/config/{company_id}", response_model=DropshipConfigResponse)
def get_dropship_config(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retorna a configuração Dropship do seller. Master ou próprio seller."""
    _require_seller_or_master(current_user, company_id)

    config = db.query(DropshipConfig).filter(
        DropshipConfig.company_id == company_id,
        DropshipConfig.provider == "ERDOS"
    ).first()

    if not config:
        # Retorna configuração vazia (não é 404 — facilita UI)
        return DropshipConfigResponse(
            id=0, company_id=company_id, provider="ERDOS", enabled=False,
            api_token=None, api_base_url="https://wxcapqbtvgttooamglxx.supabase.co/functions/v1/api-fornecedor",
            horus_customer_id=None, horus_fiscal_param_remessa=None, horus_fiscal_param_venda=None,
            stock_sync_interval_min=30, stock_sync_enabled=False, stock_sync_last_run=None,
            created_at=None, updated_at=None
        )

    # Enriquecer com dados do customer vinculado
    response = DropshipConfigResponse.model_validate(config)
    if config.horus_customer:
        response.horus_customer_name = config.horus_customer.name
        response.horus_customer_document = config.horus_customer.document
        response.horus_customer_id_guid = config.horus_customer.id_guid
        response.horus_customer_id_doc = config.horus_customer.id_doc
        response.horus_customer_cod_cli = config.horus_customer_cod_cli
    return response


@router.post("/config/{company_id}", response_model=DropshipConfigResponse)
async def upsert_dropship_config(
    company_id: int,
    payload: DropshipConfigCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Cria ou atualiza configuração Dropship. Master ou próprio seller."""
    _require_seller_or_master(current_user, company_id)

    config = db.query(DropshipConfig).filter(
        DropshipConfig.company_id == company_id,
        DropshipConfig.provider == payload.provider
    ).first()

    if config:
        for field, value in payload.model_dump(exclude_unset=True).items():
            setattr(config, field, value)
    else:
        config = DropshipConfig(company_id=company_id, **payload.model_dump())
        db.add(config)

    db.commit()
    db.refresh(config)

    # Tenta buscar e salvar o COD_CLI do Hórus se o customer estiver vinculado e ainda não tiver COD_CLI
    if config.horus_customer_id and config.horus_customer:
        doc_clean = re.sub(r"\D", "", str(config.horus_customer.document or ""))
        if doc_clean:
            try:
                from app.integrators.horus_clients import HorusClients
                horus_clients = HorusClients(db, company_id)
                busca = await horus_clients.get_client_b2c(cnpj_destino="", cpf=doc_clean)
                if busca and not busca.get("error") and busca.get("data"):
                    cod_cli = str(busca["data"].get("COD_CLI") or busca["data"].get("CODIGO") or "").strip()
                    if cod_cli:
                        config.horus_customer_cod_cli = cod_cli
                        db.commit()
                        db.refresh(config)
            except Exception as _e_save:
                log.warning(f"[Dropship] Erro ao buscar COD_CLI do customer vinculado: {_e_save}")

    response = DropshipConfigResponse.model_validate(config)
    if config.horus_customer:
        response.horus_customer_name = config.horus_customer.name
        response.horus_customer_document = config.horus_customer.document
        response.horus_customer_id_guid = config.horus_customer.id_guid
        response.horus_customer_id_doc = config.horus_customer.id_doc
        response.horus_customer_cod_cli = config.horus_customer_cod_cli
    return response


@router.post("/config/{company_id}/test-connection")
async def test_erdos_connection(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Testa conectividade com o Hub-Erdos usando as credenciais configuradas."""
    _require_seller_or_master(current_user, company_id)

    config = _get_config_or_404(db, company_id)
    client = _build_erdos_client(config)
    try:
        result = await client.test_connection()
        return {"status": "connected", "detail": result}
    except ErdosClientError as e:
        return {"status": "error", "detail": str(e)}
    finally:
        await client.close()


# ==============================================================================
# BUSCA DE CUSTOMERS (para vincular ao dropship config)
# ==============================================================================

@router.get("/config/{company_id}/customers", response_model=List[CustomerSearchResult])
def search_horus_linked_customers(
    company_id: int,
    q: Optional[str] = Query(None, description="Busca por nome ou CNPJ"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Retorna customers da empresa que possuem id_guid E id_doc configurados
    (necessário para integração Hórus). Usado para vincular o parceiro dropship.
    """
    _require_seller_or_master(current_user, company_id)

    query = db.query(Customer).filter(
        Customer.company_id == company_id,
        Customer.id_guid.isnot(None),
        Customer.id_doc.isnot(None),
        Customer.id_guid != "",
        Customer.id_doc != "",
    )

    if q:
        like = f"%{q}%"
        query = query.filter(
            (Customer.name.ilike(like)) | (Customer.document.ilike(like))
        )

    customers = query.order_by(Customer.name).limit(50).all()
    return customers


# ==============================================================================
# PEDIDOS — SINCRONIZAÇÃO
# ==============================================================================

@router.post("/orders/{company_id}/sync")
async def sync_dropship_orders(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Busca pedidos prontos para despacho no Hub-Erdos e salva/atualiza localmente.
    Baixa documentos (XML, DANFE, etiqueta) imediatamente — URLs expiram em 1h.
    """
    _require_seller_or_master(current_user, company_id)

    config = _get_config_or_404(db, company_id)
    client = _build_erdos_client(config)

    try:
        pending_orders = await client.get_pending_orders()
    except ErdosClientError as e:
        raise HTTPException(status_code=502, detail=f"Erro ao conectar com Hub-Erdos: {str(e)}")

    synced = 0
    skipped = 0
    errors = []

    base_dir = f"uploads/{company_id}/dropship"

    for order_data in pending_orders:
        external_id = order_data.get("id_pedido_erdos")
        if not external_id:
            continue

        # Verificar se já existe
        existing = db.query(DropshipOrder).filter(
            DropshipOrder.company_id == company_id,
            DropshipOrder.external_order_id == external_id
        ).first()

        # Pedido já processado (fora da fila Erdos) — não sobrescrever
        if existing and existing.status not in ["PENDING"]:
            skipped += 1
            continue

        order_dir = f"{base_dir}/{external_id}"

        # Baixar documentos imediatamente (URLs assinadas expiram em 1h)
        xml_path = None
        danfe_path = None
        label_path = None

        doc_fiscal = order_data.get("documentos_fiscais", {}) or {}
        logistica = order_data.get("logistica", {}) or {}

        xml_url = doc_fiscal.get("url_xml_nfe_venda_6120")
        danfe_url = doc_fiscal.get("url_pdf_danfe")
        label_url = logistica.get("url_pdf_etiqueta_postagem")

        try:
            if xml_url:
                xml_path = await _download_and_store(client, xml_url, f"{order_dir}/nfe.xml")
            if danfe_url:
                danfe_path = await _download_and_store(client, danfe_url, f"{order_dir}/danfe.pdf")
            if label_url:
                label_path = await _download_and_store(client, label_url, f"{order_dir}/etiqueta.pdf")
        except Exception as e:
            errors.append({"order": external_id, "error": f"Download de documentos: {str(e)}"})

        # Parsear data_liberacao
        released_at = None
        try:
            data_lib = order_data.get("data_liberacao")
            if data_lib:
                released_at = datetime.fromisoformat(data_lib.replace("Z", "+00:00"))
        except Exception:
            pass

        if existing:
            # Atualizar pedido existente (PENDING)
            existing.external_reference = order_data.get("referencia")
            existing.channel = order_data.get("canal_origem")
            existing.released_at = released_at
            existing.customer_data = order_data.get("dados_cliente")
            existing.items_data = order_data.get("itens")
            existing.logistics_data = {k: v for k, v in logistica.items() if k != "url_pdf_etiqueta_postagem"}
            existing.fiscal_data = {k: v for k, v in doc_fiscal.items() if not k.startswith("url_")}
            if xml_path:
                existing.xml_path = xml_path
            if danfe_path:
                existing.danfe_path = danfe_path
            if label_path:
                existing.label_path = label_path
            existing.synced_at = datetime.utcnow()
        else:
            new_order = DropshipOrder(
                company_id=company_id,
                config_id=config.id,
                external_order_id=external_id,
                external_reference=order_data.get("referencia"),
                channel=order_data.get("canal_origem"),
                status="PENDING",
                released_at=released_at,
                customer_data=order_data.get("dados_cliente"),
                items_data=order_data.get("itens"),
                logistics_data={k: v for k, v in logistica.items() if k != "url_pdf_etiqueta_postagem"},
                fiscal_data={k: v for k, v in doc_fiscal.items() if not k.startswith("url_")},
                xml_path=xml_path,
                danfe_path=danfe_path,
                label_path=label_path,
                synced_at=datetime.utcnow(),
            )
            db.add(new_order)
            synced += 1

    db.commit()
    await client.close()

    return {
        "synced": synced,
        "skipped": skipped,
        "errors": errors,
        "total_hub": len(pending_orders),
    }


# ==============================================================================
# PEDIDOS — LISTAGEM
# ==============================================================================

@router.get("/orders/{company_id}", response_model=List[DropshipOrderResponse])
def list_dropship_orders(
    company_id: int,
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista pedidos dropship da empresa com filtro opcional de status."""
    _require_seller_or_master(current_user, company_id)

    query = db.query(DropshipOrder).filter(DropshipOrder.company_id == company_id)
    if status:
        query = query.filter(DropshipOrder.status == status.upper())

    return query.order_by(DropshipOrder.released_at.desc().nullslast()).all()


@router.get("/orders/{company_id}/{order_id}", response_model=DropshipOrderResponse)
def get_dropship_order(
    company_id: int,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retorna detalhes de um pedido dropship específico, incluindo dados de conferência se houver."""
    _require_seller_or_master(current_user, company_id)

    order = db.query(DropshipOrder).filter(
        DropshipOrder.id == order_id,
        DropshipOrder.company_id == company_id,
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido dropship não encontrado.")

    from app.models.order_conference import OrderConference

    # Converter modelo SQLAlchemy em dicionário
    order_data = {
        col.name: getattr(order, col.name)
        for col in order.__table__.columns
    }

    from sqlalchemy import or_

    rem_raw = str(order.horus_pedido_remessa or '')
    rem_clean = rem_raw.replace('#', '').strip()
    ext_ref = str(order.external_reference or '')
    ext_clean = ext_ref.replace('#', '').strip()

    origins = list(set(filter(None, [
        rem_raw,
        rem_clean,
        f"#{rem_clean}" if rem_clean else None,
        f"RM-{ext_clean}" if ext_clean else None,
        f"RM-#{ext_clean}" if ext_clean else None,
        ext_ref,
        ext_clean,
        str(order.external_order_id or ''),
        str(order.id)
    ])))

    conf = db.query(OrderConference).filter(
        OrderConference.company_id == company_id,
        or_(
            OrderConference.cod_pedido_origem.in_(origins),
            OrderConference.cod_ped_venda.in_(origins)
        )
    ).order_by(OrderConference.updated_at.desc()).first()

    order_data["conference"] = {
        "id": conf.id,
        "branch_id": conf.branch_id,
        "status": conf.status,
        "cod_cli": conf.cod_cli,
        "cod_pedido_origem": conf.cod_pedido_origem,
        "created_at": conf.created_at.isoformat() if conf.created_at else None
    } if conf else None

    return order_data


# ==============================================================================
# HELPERS: cache de item e busca/criação de cliente final
# ==============================================================================

async def _get_horus_item_info(
    db: Session,
    company_id: int,
    isbn: str,
    horus_orders,
    cnpj_destino: str,
    id_doc_erdos: str,
    id_guid_erdos: str,
    horus_clients=None,   # opcional: se fornecido, tenta Busca_Acervo (B2C) primeiro
) -> Dict[str, Any]:
    """
    Retorna (horus_cod_item, horus_vlr_capa) para o ISBN dado.
    Usa cache local (dsp_item_cache) com TTL de 24h.

    Prioridade de busca:
    1. Cache local (TTL 24h)
    2. Busca_Acervo via horus_clients (B2C, sem CNPJ_DESTINO) — se horus_clients fornecido
    3. Busca_ProdutoB2B via horus_orders (B2B, com CNPJ_DESTINO)
    """
    from datetime import timezone, timedelta
    from app.models.dropship import DropshipItemCache

    TTL_HOURS = 24
    now = datetime.utcnow()

    cached = db.query(DropshipItemCache).filter(
        DropshipItemCache.company_id == company_id,
        DropshipItemCache.isbn == isbn,
    ).first()

    if cached and cached.horus_cod_item:
        age = now - cached.cached_at.replace(tzinfo=None) if cached.cached_at else timedelta(hours=TTL_HOURS + 1)
        if age < timedelta(hours=TTL_HOURS):
            return {"cod_item": cached.horus_cod_item, "vlr_capa": float(cached.horus_vlr_capa or 0)}

    cod_item = None
    vlr_capa = 0.0

    # 1) Busca_Acervo (B2C) — fonte primária para remessa
    if horus_clients is not None:
        try:
            acervo = await horus_clients.get(  # type: ignore[attr-defined]
                "Busca_Acervo",
                params={"BARRAS_ISBN": isbn},
            )
            if acervo and isinstance(acervo, list) and len(acervo) > 0:
                item_data = acervo[0]
                if not item_data.get("Falha"):
                    cod_item = str(item_data.get("COD_ITEM") or item_data.get("CODIGO") or "").strip() or None
                    try:
                        vlr_raw = item_data.get("VLR_CAPA") or item_data.get("PRECO") or item_data.get("VLR_LIQUIDO") or 0
                        if isinstance(vlr_raw, str):
                            vlr_raw = vlr_raw.replace(".", "").replace(",", ".")
                        vlr_capa = float(vlr_raw)
                    except (ValueError, TypeError):
                        vlr_capa = 0.0
        except Exception:
            pass

    # 2) Busca_ProdutoB2B (B2B) — fallback
    if not cod_item:
        try:
            result = await horus_orders.get(  # type: ignore[attr-defined]
                "Busca_ProdutoB2B",
                params={
                    "ID_DOC": id_doc_erdos,
                    "ID_GUID": id_guid_erdos,
                    "CNPJ_DESTINO": cnpj_destino,
                    "BARRAS_ISBN": isbn,
                    "LIMIT": 1,
                    "OFFSET": 0,
                }
            )
        except Exception:
            result = None

        if result and isinstance(result, list) and len(result) > 0:
            item_data = result[0]
            if not item_data.get("Falha"):
                cod_item = str(item_data.get("COD_ITEM") or item_data.get("CODIGO") or "").strip() or None
                try:
                    vlr_raw = item_data.get("VLR_CAPA") or item_data.get("PRECO") or item_data.get("VLR_LIQUIDO") or 0
                    if isinstance(vlr_raw, str):
                        vlr_raw = vlr_raw.replace(".", "").replace(",", ".")
                    vlr_capa = float(vlr_raw)
                except (ValueError, TypeError):
                    vlr_capa = 0.0


    # Salvar/atualizar cache
    if cached:
        cached.horus_cod_item = cod_item
        cached.horus_vlr_capa = vlr_capa
        cached.cached_at = now
    else:
        db.add(DropshipItemCache(
            company_id=company_id,
            isbn=isbn,
            horus_cod_item=cod_item,
            horus_vlr_capa=vlr_capa,
            cached_at=now,
        ))
    try:
        db.commit()
    except Exception:
        db.rollback()

    return {"cod_item": cod_item, "vlr_capa": vlr_capa}


async def _get_or_create_horus_customer(
    horus_clients,
    customer_data: Dict[str, Any],
    config,
    cnpj_destino: str,
    seller_company=None,         # Company model — usado como fallback de endereço
) -> Optional[str]:
    """
    Busca o cliente final     - CPF ausente → retorna None (quem chama decide o que fazer)
    - Endereço ausente → usa dados do seller como fallback
    - Falha no Hórus → levanta exceção com mensagem clara
    """
    import logging
    log = logging.getLogger(__name__)

    cpf_cnpj = re.sub(r"\D", "", str(customer_data.get("cpf_cnpj") or ""))
    if not cpf_cnpj:
        return None  # Chamador deve tratar como erro crítico

    # ── Busca pelo CPF/CNPJ no Hórus ──────────────────────────────────────────
    busca_erro: Optional[str] = None
    try:
        busca = await horus_clients.get_client_b2c(  # type: ignore[attr-defined]
            cnpj_destino=cnpj_destino,
            cpf=cpf_cnpj,
        )
        if busca and not busca.get("error") and busca.get("data"):
            data = busca["data"]
            cod_cli = str(data.get("COD_CLI") or data.get("CODIGO") or "").strip()
            if cod_cli:
                log.info(f"[Dropship] Cliente encontrado no Hórus: CPF={cpf_cnpj} COD_CLI={cod_cli}")
                # Atualiza endereço do cliente encontrado (mesmo fluxo do InsAltEndCliente)
                # O endereço é atualizado assincronamente — não bloqueia
                try:
                    _UF_NOMES_QUICK = {
                        "AC": "Acre", "AL": "Alagoas", "AP": "Amapá", "AM": "Amazonas",
                        "BA": "Bahia", "CE": "Ceará", "DF": "Distrito Federal",
                        "ES": "Espírito Santo", "GO": "Goiás", "MA": "Maranhão",
                        "MT": "Mato Grosso", "MS": "Mato Grosso do Sul", "MG": "Minas Gerais",
                        "PA": "Pará", "PB": "Paraíba", "PR": "Paraná", "PE": "Pernambuco",
                        "PI": "Piauí", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
                        "RS": "Rio Grande do Sul", "RO": "Rondônia", "RR": "Roraima",
                        "SC": "Santa Catarina", "SP": "São Paulo", "SE": "Sergipe", "TO": "Tocantins",
                    }
                    _uf  = str(customer_data.get("uf") or "").strip().upper()
                    _cep = re.sub(r"\D", "", str(customer_data.get("cep") or ""))
                    _end_params: Dict[str, Any] = {
                        "COD_CLI":       cod_cli,
                        "COD_TPO_END":   config.horus_cod_endereco or "1",
                        "NOM_PAIS":      "Brasil",
                        "SIGLA_UF":      _uf,
                        "NOME_UF":       _UF_NOMES_QUICK.get(_uf, _uf),
                        "NOM_LOCAL":     str(customer_data.get("cidade") or "")[:60],
                        "NOM_BAIRRO":    str(customer_data.get("bairro") or "")[:40],
                        "DESC_ENDERECO": str(customer_data.get("endereco") or "")[:80],
                        "NUM_END":       str(customer_data.get("numero") or "S/N")[:10],
                        "CEP":           _cep,
                    }
                    await horus_clients.get("InsAltEndCliente", params=_end_params)  # type: ignore[attr-defined]
                    log.info(f"[Dropship] Endereço atualizado para COD_CLI={cod_cli}")
                except Exception as _ex:
                    log.warning(f"[Dropship] InsAltEndCliente (atualização) não bloqueante: {_ex}")
                return cod_cli
        # Não encontrado — vai tentar criar
        busca_erro = busca.get("msg") if busca else None
        log.info(f"[Dropship] Cliente CPF={cpf_cnpj} não encontrado no Hórus — tentando criar. msg={busca_erro}")
    except Exception as e:
        log.warning(f"[Dropship] Busca_ClienteB2B falhou (CPF={cpf_cnpj}): {e}")

    # ── Monta dados para InsAltCliente / InsAltEndCliente ──────────────────────
    nome     = str(customer_data.get("nome") or "").strip()
    cep      = re.sub(r"\D", "", str(customer_data.get("cep") or ""))
    uf       = str(customer_data.get("uf") or "").strip().upper()
    cidade   = str(customer_data.get("cidade") or "").strip()
    bairro   = str(customer_data.get("bairro") or "").strip()
    endereco = str(customer_data.get("endereco") or "").strip()
    numero   = str(customer_data.get("numero") or "S/N").strip()

    # Fallback de endereço: se o cliente não tem endereço, usa dados do seller
    if seller_company and (not cep or not uf or not cidade):
        cep      = cep or re.sub(r"\D", "", str(getattr(seller_company, "zip_code", "") or ""))
        uf       = uf or str(getattr(seller_company, "state", "") or "").strip().upper()
        cidade   = cidade or str(getattr(seller_company, "city", "") or "").strip()
        bairro   = bairro or str(getattr(seller_company, "neighborhood", "") or "").strip()
        endereco = endereco or str(getattr(seller_company, "address", "") or "").strip()

    doc_key    = "CPF" if len(cpf_cnpj) == 11 else "CNPJ"
    tpo_pessoa = "F"  if len(cpf_cnpj) == 11 else "J"

    # ── PASSO 1: InsAltCliente — dados cadastrais básicos (SEM endereço) ───────
    ins_params: Dict[str, Any] = {
        "COD_CLI":    "NOVO",       # obrigatório para inserção de novo cliente
        "TPO_PESSOA": tpo_pessoa,   # F=Física, J=Jurídica
        doc_key:      cpf_cnpj,
        "NOM_CLI":    nome[:60] if nome else "CLIENTE DROPSHIP",
    }

    # Parâmetros opcionais da config do seller
    if config.horus_resp_cliente:
        ins_params["NOM_RESP"] = config.horus_resp_cliente
    if config.horus_cod_resp:
        ins_params["COD_RESPONSAVEL"] = config.horus_cod_resp
    # TIPO_CLI não é parâmetro do InsAltCliente (B2C) — não enviar

    log.info(f"[Dropship] InsAltCliente params: {ins_params}")

    try:
        ins_result = await horus_clients.get(  # type: ignore[attr-defined]
            "InsAltCliente", params=ins_params
        )
        log.info(f"[Dropship] InsAltCliente resposta: {ins_result}")

        if ins_result and isinstance(ins_result, list) and len(ins_result) > 0:
            item = ins_result[0]
            if item.get("Falha"):
                msg_horus = item.get("Mensagem") or item.get("MSG") or "Erro desconhecido no Hórus"
                # Caso especial: duplicação — cliente já existe, tentar buscar o COD_CLI
                if "duplica" in msg_horus.lower() or "buscar antes" in msg_horus.lower():
                    log.warning(f"[Dropship] InsAltCliente: cliente já existe no Hórus — buscando COD_CLI existente")
                    busca_retry = await horus_clients.get_client_b2c(  # type: ignore[attr-defined]
                        cnpj_destino=cnpj_destino, cpf=cpf_cnpj
                    )
                    if busca_retry and not busca_retry.get("error") and busca_retry.get("data"):
                        cod_cli_retry = str(busca_retry["data"].get("COD_CLI") or busca_retry["data"].get("CODIGO") or "").strip()
                        if cod_cli_retry:
                            log.info(f"[Dropship] COD_CLI recuperado após duplicação: {cod_cli_retry}")
                            return cod_cli_retry
                raise ValueError(f"InsCliente retornou Falha: {msg_horus}")

            cod_cli = str(item.get("COD_CLI") or item.get("CODIGO") or "").strip()
            if not cod_cli:
                raise ValueError(f"InsAltCliente não retornou COD_CLI. Resposta: {item}")
        else:
            raise ValueError(f"InsAltCliente retornou resposta vazia ou inválida: {ins_result}")

    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Falha na comunicação com Hórus ao criar cliente (InsAltCliente): {e}") from e

    log.info(f"[Dropship] Cliente criado com sucesso: COD_CLI={cod_cli}")

    # ── PASSO 2: InsAltEndCliente — cadastrar endereço do cliente ──────────────
    # Mapeamento de UF → nome completo do estado (obrigatório pelo Hórus)
    _UF_NOMES = {
        "AC": "Acre", "AL": "Alagoas", "AP": "Amapá", "AM": "Amazonas",
        "BA": "Bahia", "CE": "Ceará", "DF": "Distrito Federal",
        "ES": "Espírito Santo", "GO": "Goiás", "MA": "Maranhão",
        "MT": "Mato Grosso", "MS": "Mato Grosso do Sul", "MG": "Minas Gerais",
        "PA": "Pará", "PB": "Paraíba", "PR": "Paraná", "PE": "Pernambuco",
        "PI": "Piauí", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
        "RS": "Rio Grande do Sul", "RO": "Rondônia", "RR": "Roraima",
        "SC": "Santa Catarina", "SP": "São Paulo", "SE": "Sergipe", "TO": "Tocantins",
    }
    nome_uf = _UF_NOMES.get(uf, uf)

    end_params: Dict[str, Any] = {
        "COD_CLI":      cod_cli,
        "NOM_PAIS":     "Brasil",
        "SIGLA_UF":     uf,
        "NOME_UF":      nome_uf,
        "NOM_LOCAL":    cidade[:60] if cidade else "",
        "NOM_BAIRRO":   bairro[:40] if bairro else "",
        "DESC_ENDERECO": endereco[:80] if endereco else "",
        "NUM_END":      numero[:10],
        "CEP":          cep,
    }

    # Tipo de endereço: vem da config ou padrão 1
    end_params["COD_TPO_END"] = config.horus_cod_endereco or "1"

    if customer_data.get("complemento"):
        end_params["COM_ENDERECO"] = str(customer_data["complemento"])[:40]

    log.info(f"[Dropship] InsAltEndCliente params: {end_params}")

    try:
        end_result = await horus_clients.get(  # type: ignore[attr-defined]
            "InsAltEndCliente", params=end_params
        )
        log.info(f"[Dropship] InsAltEndCliente resposta: {end_result}")

        if end_result and isinstance(end_result, list) and len(end_result) > 0:
            item_end = end_result[0]
            if item_end.get("Falha"):
                msg = item_end.get("Mensagem") or "Erro ao cadastrar endereço no Hórus"
                log.warning(f"[Dropship] InsAltEndCliente Falha (não bloqueante): {msg}")
                # Endereço não bloqueante — cliente já foi criado, pedido pode continuar
        else:
            log.warning(f"[Dropship] InsAltEndCliente retornou vazio: {end_result}")

    except Exception as e:
        # Endereço não bloqueia o pedido — apenas loga
        log.warning(f"[Dropship] InsAltEndCliente falhou (não bloqueante): {e}")

    return cod_cli



def _validate_preflight(
    order,
    config,
    customer_data: Dict[str, Any],
    fiscal_intra: Optional[str],
    fiscal_inter: Optional[str],
) -> List[str]:
    """
    Valida todos os dados obrigatórios ANTES de qualquer chamada ao Hórus.
    Retorna lista de erros (vazia = tudo OK).

    Parâmetros obrigatórios (InsPedidoVenda com TIPO_PEDIDO_V_T_D='L'):
    - CPF/CNPJ do cliente final (necessário para InsCliente/Busca_ClienteB2B)
    - COD_PARAM_FISCAL (intra ou inter)
    - COD_PARAM_FISCAL da Venda
    - Customer ERDOS vinculado
    - Pedido deve estar em status PENDING
    - Pedido não pode já ter sido enviado (idempotência)

    NÃO obrigatórios por definição do fluxo L (Logística):
    - COD_FORMA, QTD_PARCELAS, CONDICAO_PAGAMENTO
    """
    erros: List[str] = []

    # --- Idempotência: já foi enviado? ---
    if order.horus_pedido_remessa or order.horus_pedido_venda:
        erros.append(
            f"Pedido já foi enviado ao Hórus anteriormente "
            f"(Remessa: {order.horus_pedido_remessa or '-'} | "
            f"Venda: {order.horus_pedido_venda or '-'}). "
            f"Não é possível reenviar para evitar duplicidade."
        )
        return erros  # retorna imediatamente — sem checar mais nada

    # --- Status do pedido ---
    if order.status != "PENDING":
        erros.append(f"Pedido não está pendente (status: {order.status}). Apenas pedidos PENDING podem ser enviados.")

    # --- Itens ---
    itens = order.items_data or []
    if not itens:
        erros.append("Pedido sem itens. É necessário ao menos 1 item para envio.")
    else:
        for i, item in enumerate(itens, 1):
            isbn = (
                item.get("sku") or item.get("sku_fornecedor")
                or item.get("isbn") or item.get("barras_isbn") or ""
            ).strip()
            if not isbn:
                erros.append(f"Item #{i} sem SKU/ISBN — campo obrigatório para identificação no Hórus.")

    # --- CPF/CNPJ do cliente final (obrigatório para InsCliente) ---
    cpf_cnpj = re.sub(r"\D", "", str(customer_data.get("cpf_cnpj") or ""))
    if not cpf_cnpj:
        erros.append(
            "CPF/CNPJ do cliente final é obrigatório para registrar o pedido de Remessa no Hórus. "
            "O pedido Erdos não contém o documento do consumidor — "
            "verifique os dados recebidos do Hub-Erdos."
        )

    # --- Paramétros fiscais da Remessa ---
    if not fiscal_intra and not fiscal_inter:
        erros.append(
            "Parâmetro fiscal de REMESSA não configurado (intraestadual e interestadual). "
            "Acesse Configurações → Dropship Horus e preencha os campos."
        )

    # --- Parâmetro fiscal da Venda ---
    if not config.horus_fiscal_param_venda:
        erros.append(
            "Parâmetro fiscal de VENDA não configurado. "
            "Acesse Configurações → Dropship Horus e preencha o campo COD_PARAM_FISCAL Venda."
        )

    # --- Customer ERDOS ---
    if not config.horus_customer_id:
        erros.append("Customer parceiro (ERDOS) não vinculado na configuração Dropship.")

    return erros


# ==============================================================================
# ENVIAR PARA O HÓRUS (Gera 2 pedidos)
# ==============================================================================

@router.post("/orders/{company_id}/{order_id}/send-to-horus")
async def send_order_to_horus(
    company_id: int,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Envia o pedido Dropship para o Hórus ERP criando dois pedidos:

    1. Pedido de Remessa (CFOP 6.923) — com dados do cliente CONSUMIDOR FINAL
       - Busca cliente por CPF no Hórus; se não encontrar, cria via InsCliente
       - COD_ITEM obtido via Busca_ProdutoB2B (cacheado por 24h por ISBN)
       - COD_PARAM_FISCAL: intraestadual ou interestadual (por UF)

    2. Pedido de Venda (CFOP 6.118) — com customer ERDOS configurado (inalterado)
       - COD_ITEM e VLR_CAPA do mesmo cache da remessa

    Ambos terminam em status LEX.
    """
    _require_seller_or_master(current_user, company_id)

    # 1. Carregar pedido dropship
    order = db.query(DropshipOrder).filter(
        DropshipOrder.id == order_id,
        DropshipOrder.company_id == company_id,
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido dropship não encontrado.")

    # 2. Carregar configuração dropship
    config = db.query(DropshipConfig).filter(
        DropshipConfig.id == order.config_id
    ).first()
    if not config:
        raise HTTPException(status_code=400, detail="Configuração Dropship não encontrada.")

    fiscal_intra = config.horus_fiscal_param_remessa_intra or config.horus_fiscal_param_remessa
    fiscal_inter = config.horus_fiscal_param_remessa_inter or config.horus_fiscal_param_remessa

    # 3. Carregar empresa e customer ERDOS
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Empresa não encontrada.")

    customer_erdos = db.query(Customer).filter(Customer.id == config.horus_customer_id).first() if config.horus_customer_id else None
    if not customer_erdos or not customer_erdos.id_guid or not customer_erdos.id_doc:
        raise HTTPException(status_code=400, detail="Customer ERDOS não possui id_guid/id_doc configurados.")

    # 4. Validação Pré-Voo (✅ tudo antes de qualquer chamada ao Hórus)
    customer_data = order.customer_data or {}
    preflight_errors = _validate_preflight(
        order=order,
        config=config,
        customer_data=customer_data,
        fiscal_intra=fiscal_intra,
        fiscal_inter=fiscal_inter,
    )
    if preflight_errors:
        raise HTTPException(
            status_code=422,
            detail={
                "mensagem": "Validação falhou. Corrija os erros abaixo antes de reenviar o pedido ao Hórus.",
                "erros": preflight_errors,
            }
        )

    # 5. Verificar integração Hórus habilitada
    settings = db.query(CompanySettings).filter(CompanySettings.company_id == company_id).first()
    if not settings or not settings.horus_enabled:
        raise HTTPException(status_code=400, detail="Integração Hórus não habilitada.")

    from app.integrators.horus_clients import HorusClients
    from app.integrators.horus_orders import HorusOrders

    # 5. Preparar dados base
    customer_data = order.customer_data or {}
    ext_id        = str(order.external_reference or order.external_order_id or order.id)
    cod_origem_remessa = f"RM-{ext_id[:20]}"
    cod_origem_venda   = f"VD-{ext_id[:20]}"
    cnpj_destino       = re.sub(r"\D", "", str(company.document or ""))
    id_doc_erdos       = re.sub(r"\D", "", str(customer_erdos.id_doc)) if customer_erdos.id_doc else ""
    id_guid_erdos      = customer_erdos.id_guid or ""

    # 6. Determinar parâmetro fiscal da remessa (intra × inter estado)
    uf_seller  = str(getattr(settings, "state", "") or getattr(company, "state", "") or "").strip().upper()
    uf_cliente = str(customer_data.get("uf") or "").strip().upper()
    if uf_seller and uf_cliente and uf_seller == uf_cliente:
        fiscal_remessa = fiscal_intra or fiscal_inter
    else:
        fiscal_remessa = fiscal_inter or fiscal_intra

    horus_clients = HorusClients(db, company_id)
    horus_orders  = HorusOrders(db, company_id)

    errors: list = []
    cod_ped_remessa = None
    cod_ped_venda   = None
    cod_cli_final   = order.horus_cod_cli_final  # reutiliza se já foi buscado antes

    # Buscar COD_CLI do cliente parceiro ERDOS no Hórus (reutiliza salvo em config ou busca e salva)
    cod_cli_erdos = getattr(config, "horus_customer_cod_cli", None)
    if not cod_cli_erdos and customer_erdos and customer_erdos.document:
        doc_erdos_clean = re.sub(r"\D", "", str(customer_erdos.document))
        try:
            busca_erdos = await horus_clients.get_client_b2c(cnpj_destino="", cpf=doc_erdos_clean)
            if busca_erdos and not busca_erdos.get("error") and busca_erdos.get("data"):
                cod_cli_erdos = str(busca_erdos["data"].get("COD_CLI") or busca_erdos["data"].get("CODIGO") or "").strip()
                if cod_cli_erdos and config:
                    config.horus_customer_cod_cli = cod_cli_erdos
                    db.commit()
                    log.info(f"[Dropship] COD_CLI do cliente ERDOS no Hórus salvo na config: {cod_cli_erdos}")
        except Exception as _e_erdos:
            log.warning(f"[Dropship] Busca_Cliente p/ Erdos falhou: {_e_erdos}")

    log.info(f"[Dropship] COD_CLI ERDOS confirmado p/ alteração de status B2B: {cod_cli_erdos}")

    try:
        itens = order.items_data or []

        # ───────────────────────────────────────────────────────────────────
        # STEP A — Pré-buscar COD_ITEM e VLR_CAPA de todos os itens (cache)
        # ───────────────────────────────────────────────────────────────────
        item_info_map: Dict[str, Dict[str, Any]] = {}
        for item in itens:
            isbn = (
                item.get("sku")
                or item.get("sku_fornecedor")
                or item.get("isbn")
                or item.get("barras_isbn")
                or ""
            ).strip()
            if isbn and isbn not in item_info_map:
                info = await _get_horus_item_info(
                    db=db,
                    company_id=company_id,
                    isbn=isbn,
                    horus_orders=horus_orders,
                    cnpj_destino=cnpj_destino,
                    id_doc_erdos=id_doc_erdos,
                    id_guid_erdos=id_guid_erdos,
                    horus_clients=horus_clients,   # habilita Busca_Acervo (B2C) como fonte primária
                )
                item_info_map[isbn] = info

        # ───────────────────────────────────────────────────────────────────
        # STEP B — Buscar/criar cliente final no Hórus
        # ───────────────────────────────────────────────────────────────────
        if not cod_cli_final:
            try:
                cod_cli_final = await _get_or_create_horus_customer(
                    horus_clients=horus_clients,
                    customer_data=customer_data,
                    config=config,
                    cnpj_destino=cnpj_destino,
                    seller_company=company,  # fallback de endereço
                )
            except ValueError as e:
                raise HTTPException(
                    status_code=422,
                    detail={
                        "mensagem": "Falha ao registrar o cliente no Hórus. Nenhum pedido foi enviado.",
                        "erros": [str(e)],
                    }
                )

        # Segurança final: se ainda None, CPF ausente (não deve chegar aqui após preflight)
        if not cod_cli_final:
            raise HTTPException(
                status_code=422,
                detail={
                    "mensagem": "Não foi possível registrar ou localizar o cliente no Hórus.",
                    "erros": [
                        "CPF/CNPJ não retornou COD_CLI válido. Verifique os dados e tente novamente."
                    ]
                }
            )


        # ───────────────────────────────────────────────────────────────────
        # PEDIDO 1 — REMESSA (CFOP 6.923): B2C — entrega ao cliente final
        # Usa horus_clients (sem CNPJ_DESTINO) + parâmetros B2C obrigatórios
        # ───────────────────────────────────────────────────────────────────
        remessa_params: Dict[str, Any] = {
            "TIPO_PEDIDO_V_T_D":  "L",
            "COD_PARAM_FISCAL":   fiscal_remessa,
            "COD_PEDIDO_ORIGEM":  cod_origem_remessa,
            "OBS_PEDIDO": (
                f"Dropship Erdos | Pedido Erdos #{order.external_order_id} | Ref: {order.external_reference or order.external_order_id} "
                f"| Cliente: {customer_data.get('nome', 'N/A')} "
                f"| CPF: {customer_data.get('cpf_cnpj', 'N/A')} "
                f"| CEP: {customer_data.get('cep', 'N/A')}"
            ),
        }
        # COD_EMPRESA e COD_FILIAL — obrigatórios B2C
        if settings.horus_company:
            remessa_params["COD_EMPRESA"] = settings.horus_company
        if settings.horus_branch:
            remessa_params["COD_FILIAL"] = settings.horus_branch
        # COD_CLI — cliente final encontrado/criado (obrigatório B2C)
        if cod_cli_final:
            remessa_params["COD_CLI"] = cod_cli_final
        # Parâmetros opcionais/obrigatórios B2C da config do seller
        if config.horus_cod_metodo:
            remessa_params["COD_METODO"] = config.horus_cod_metodo
        if config.horus_cod_endereco:
            remessa_params["COD_TPO_END"] = config.horus_cod_endereco
        if config.horus_cod_transp:
            remessa_params["COD_TRANSP"] = config.horus_cod_transp
        if config.horus_frete_emit_dest:
            remessa_params["FRETE_EMIT_DEST"] = config.horus_frete_emit_dest
        # NÃO enviar: COD_FORMA, QTD_PARCELAS, CONDICAO_PAGAMENTO, CNPJ_DESTINO, ID_DOC, ID_GUID

        remessa_result = await horus_clients.get(  # type: ignore[attr-defined]
            "InsPedidoVenda", params=remessa_params
        )

        if remessa_result and isinstance(remessa_result, list):
            if remessa_result[0].get("Falha"):
                # Remessa falhou — ABORTAR COMPLETAMENTE, não enviar Venda
                msg_horus = remessa_result[0].get('Mensagem', 'Erro desconhecido no Hórus')
                raise HTTPException(
                    status_code=422,
                    detail={
                        "mensagem": "Falha ao criar o Pedido de Remessa no Hórus. O Pedido de Venda NÃO foi enviado.",
                        "erros": [f"Remessa (InsPedidoVenda): {msg_horus}"],
                    }
                )
            else:
                cod_ped_remessa = remessa_result[0].get("COD_PED_VENDA")

        if not cod_ped_remessa:
            raise HTTPException(
                status_code=422,
                detail={
                    "mensagem": "Hórus não retornou o código do Pedido de Remessa. O Pedido de Venda NÃO foi enviado.",
                    "erros": ["Verifique os parâmetros enviados e tente novamente."],
                }
            )

        # Remessa criada — inserir itens (B2C)
        for item in itens:
            isbn = (
                item.get("sku") or item.get("sku_fornecedor")
                or item.get("isbn") or item.get("barras_isbn") or ""
            ).strip()
            qty = int(item.get("quantidade") or item.get("qty") or 1)
            if not isbn:
                errors.append(f"Item sem SKU/ISBN: {item}")
                continue

            info = item_info_map.get(isbn, {})
            cod_item = info.get("cod_item")
            vlr_capa = info.get("vlr_capa", 0.0)

            if not cod_item:
                errors.append(f"COD_ITEM não encontrado para ISBN={isbn} — item não inserido na remessa")
                continue

            # Parâmetros B2C obrigatórios para InsItensPedidoVenda
            # Aplicar desconto configurado sobre VLR_LIQUIDO (vlr_capa)
            perc_desc = float(config.perc_desconto_remessa or 0)
            vlr_liq = round(vlr_capa * (1 - perc_desc / 100), 2) if perc_desc > 0 else float(vlr_capa)
            item_params_remessa: Dict[str, Any] = {
                "COD_PED_VENDA": cod_ped_remessa,
                "COD_ITEM":      cod_item,
                "QTD_PEDIDA":    qty,
                "VLR_LIQUIDO":   vlr_liq,
            }
            if settings.horus_company:
                item_params_remessa["COD_EMPRESA"] = settings.horus_company
            if settings.horus_branch:
                item_params_remessa["COD_FILIAL"] = settings.horus_branch
            if cod_cli_final:
                item_params_remessa["COD_CLI"] = cod_cli_final

            await horus_clients.get(  # type: ignore[attr-defined]
                "InsItensPedidoVenda", params=item_params_remessa
            )

        # Muda status Remessa — usando a especificação Hórus B2C (AltStatus_Pedido)
        if cod_ped_remessa:
            target_status = (config.horus_status_envio_erp or "LEX").strip()
            alt_status_params: Dict[str, Any] = {
                "COD_EMPRESA":   settings.horus_company,
                "COD_FILIAL":    settings.horus_branch,
                "COD_CLI":       cod_cli_final,
                "COD_PED_VENDA": cod_ped_remessa,
                "STA_PEDIDO":    target_status,
            }

            # Remove chaves None/Vazias
            alt_status_params = {k: v for k, v in alt_status_params.items() if v is not None and v != ""}

            try:
                alt_res = await horus_clients.get(  # type: ignore[attr-defined]
                    "AltStatus_Pedido", params=alt_status_params
                )
                log.info(f"[Dropship] AltStatus_Pedido remessa resposta: {alt_res}")
                if alt_res and isinstance(alt_res, list) and len(alt_res) > 0:
                    item_alt = alt_res[0]
                    if item_alt.get("Falha") or item_alt.get("FALHA") == "S":
                        msg_alt = item_alt.get("Mensagem") or item_alt.get("MENSAGEM") or str(item_alt)
                        errors.append(f"AltStatus_Pedido Remessa ({target_status}): {msg_alt}")
                elif alt_res and isinstance(alt_res, dict):
                    if alt_res.get("Falha") or alt_res.get("FALHA") == "S":
                        msg_alt = alt_res.get("Mensagem") or alt_res.get("MENSAGEM") or str(alt_res)
                        errors.append(f"AltStatus_Pedido Remessa ({target_status}): {msg_alt}")
            except Exception as _e:
                log.error(f"[Dropship] AltStatus_Pedido remessa falhou: {_e}")
                errors.append(f"AltStatus_Pedido Remessa ({target_status}): {_e}")

        # ─────────────────────────────────────────────────────────────────────────
        # PEDIDO 2 — VENDA (CFOP 6.118): customer ERDOS → LAP → LFT
        # SEGURANÇA: só chegamos aqui se a Remessa foi criada com sucesso.
        # ─────────────────────────────────────────────────────────────────────────
        # Enviado como padrão B2B com TIPO_PEDIDO_V_T_D='L' (Logística).
        # O COD_PARAM_FISCAL é pré-configurado nas settings do cliente ERDOS no Hórus.
        # NÃO enviar COD_FORMA, QTD_PARCELAS, CONDICAO_PAGAMENTO neste pedido.
        # Fluxo de status: [criado] → AltStatus (LAP) → Pular_expedicao (LFT)
        venda_params: Dict[str, Any] = {
            "ID_DOC":            id_doc_erdos,
            "ID_GUID":           id_guid_erdos,
            "CNPJ_DESTINO":      cnpj_destino,
            "TIPO_PEDIDO_V_T_D": "L",
            "COD_PEDIDO_ORIGEM": cod_origem_venda,
            "COD_PARAM_FISCAL":  config.horus_fiscal_param_venda,
            "OBS_PEDIDO": (
                f"Dropship Venda Erdos | Pedido Erdos #{order.external_order_id} | Ref: {order.external_reference or order.external_order_id}"
            ),
            # NOTA: NÃO enviar COD_FORMA, QTD_PARCELAS, CONDICAO_PAGAMENTO
        }
        # Adicionar VLR_FRETE ao pedido de Venda se configurado
        taxa_frete = float(config.vlr_taxa_frete or 0)
        if taxa_frete > 0:
            venda_params["VLR_FRETE"] = round(taxa_frete, 2)
        # COD_EMPRESA, COD_FILIAL e COD_METODO NÃO devem ser enviados no pedido de Venda B2B

        venda_result = await horus_orders.get(  # type: ignore[attr-defined]
            "InsPedidoVenda", params=venda_params
        )

        if venda_result and isinstance(venda_result, list):
            if venda_result[0].get("Falha"):
                errors.append(f"Venda: {venda_result[0].get('Mensagem', 'Erro desconhecido')}")
            else:
                cod_ped_venda = venda_result[0].get("COD_PED_VENDA")

        if cod_ped_venda:
            for item in itens:
                isbn = (
                    item.get("sku") or item.get("sku_fornecedor")
                    or item.get("isbn") or item.get("barras_isbn") or ""
                ).strip()
                qty = int(item.get("quantidade") or item.get("qty") or 1)
                if not isbn:
                    continue

                info = item_info_map.get(isbn, {})
                vlr_capa = info.get("vlr_capa", 0.0)

                # ── Tabela de Preços: desconto por ISBN (Venda 6.118 apenas) ──
                # Consulta dsp_price_table — se ISBN existe e validade futura,
                # aplica desconto e envia VLR_LIQUIDO. Caso contrário, não envia.
                price_entry = db.query(DropshipPriceTable).filter(
                    DropshipPriceTable.company_id == company_id,
                    DropshipPriceTable.isbn == isbn,
                    DropshipPriceTable.data_validade >= date.today(),
                ).first()

                if price_entry and vlr_capa:
                    desc_pct = float(price_entry.desconto or 0)
                    vlr_liquido_venda = round(float(vlr_capa) * (1 - desc_pct / 100), 2)
                    price_to_send: Optional[float] = vlr_liquido_venda
                    log.info(
                        f"[Dropship][Venda][PriceTable] ISBN={isbn} desconto={desc_pct}% "
                        f"vlr_capa={vlr_capa} → VLR_LIQUIDO={vlr_liquido_venda}"
                    )
                else:
                    price_to_send = None  # sem desconto: não envia VLR_LIQUIDO

                res_item = await horus_orders.send_order_item(  # type: ignore[attr-defined]
                    id_doc=id_doc_erdos,
                    id_guid=id_guid_erdos,
                    cnpj_destino=cnpj_destino,
                    cod_pedido_origem=cod_origem_venda,
                    isbn=isbn,
                    qty=qty,
                    price=price_to_send,  # VLR_LIQUIDO só se ISBN na tabela de preços e valido
                )
                log.info(f"[Dropship] InsItensPedidoVenda (Venda B2B) resposta p/ ISBN={isbn}: {res_item}")
                if res_item and isinstance(res_item, list) and len(res_item) > 0:
                    msg_item = res_item[0].get("Mensagem") or res_item[0].get("MENSAGEM") or ""
                    if res_item[0].get("Falha") or (msg_item and msg_item != "REGISTRO ENVIADO COM SUCESSO!"):
                        errors.append(f"Venda item {isbn}: {msg_item or 'Falha ao inserir item'}")

            # STEP 1: AltStatus → LAP (COD_EMPRESA, COD_FILIAL, COD_CLI, COD_PED_VENDA, STA_PEDIDO)
            try:
                alt_venda_params: Dict[str, Any] = {
                    "COD_PED_VENDA": cod_ped_venda,
                    "STA_PEDIDO":    "LAP",
                }
                if settings.horus_company:
                    alt_venda_params["COD_EMPRESA"] = settings.horus_company
                if settings.horus_branch:
                    alt_venda_params["COD_FILIAL"] = settings.horus_branch
                if cod_cli_erdos:
                    alt_venda_params["COD_CLI"] = cod_cli_erdos

                res_alt_venda = await horus_clients.get(  # type: ignore[attr-defined]
                    "AltStatus_Pedido", params=alt_venda_params
                )
                log.info(f"[Dropship] AltStatus_Pedido venda B2B resposta: {res_alt_venda}")
                if res_alt_venda and isinstance(res_alt_venda, list) and len(res_alt_venda) > 0:
                    item_alt = res_alt_venda[0]
                    if item_alt.get("Falha") or item_alt.get("FALHA") == "S":
                        msg_alt = item_alt.get("Mensagem") or item_alt.get("MENSAGEM") or str(item_alt)
                        errors.append(f"Venda AltStatus (LAP): {msg_alt}")
            except Exception as e_lap:
                errors.append(f"Venda AltStatus (LAP): {str(e_lap)}")

            # STEP 2: Pular_expedicao → LFT (COD_EMPRESA, COD_FILIAL, COD_CLI, COD_PED_VENDA, COD_LOCAL)
            try:
                cod_local_val = getattr(settings, 'horus_stock_local', None)
                if not cod_local_val or str(cod_local_val).strip() == '':
                    cod_local_val = 0

                pular_params: Dict[str, Any] = {
                    "COD_PED_VENDA": cod_ped_venda,
                    "COD_LOCAL":     cod_local_val,
                }
                if settings.horus_company:
                    pular_params["COD_EMPRESA"] = settings.horus_company
                if settings.horus_branch:
                    pular_params["COD_FILIAL"] = settings.horus_branch
                if cod_cli_erdos:
                    pular_params["COD_CLI"] = cod_cli_erdos

                res_pular = await horus_clients.get(  # type: ignore[attr-defined]
                    "Pular_expedicao", params=pular_params
                )
                log.info(f"[Dropship] Pular_expedicao venda B2B resposta: {res_pular}")
                if res_pular and isinstance(res_pular, list) and len(res_pular) > 0:
                    item_p = res_pular[0]
                    if item_p.get("Falha") or item_p.get("FALHA") == "S":
                        msg_p = item_p.get("Mensagem") or item_p.get("MENSAGEM") or str(item_p)
                        errors.append(f"Venda Pular_expedicao (LFT): {msg_p}")
            except Exception as e_pular:
                errors.append(f"Venda Pular_expedicao (LFT): {str(e_pular)}")


    except Exception as e:
        await horus_orders.close()  # type: ignore[attr-defined]
        await horus_clients.close()  # type: ignore[attr-defined]
        raise HTTPException(status_code=500, detail=f"Erro na integração Hórus: {str(e)}")

    await horus_orders.close()   # type: ignore[attr-defined]
    await horus_clients.close()  # type: ignore[attr-defined]

    # Persistir resultados no pedido
    if cod_ped_remessa or cod_ped_venda:
        order.status = "SENT_TO_HORUS"
        order.horus_pedido_remessa = cod_ped_remessa
        order.horus_pedido_venda   = cod_ped_venda
        order.horus_cod_cli_final  = cod_cli_final
        order.sent_to_horus_at     = datetime.utcnow()
        db.commit()

        # Notificar Erdos: muda status para "preparando"
        try:
            config_obj = db.query(DropshipConfig).filter(DropshipConfig.id == order.config_id).first()
            if config_obj:
                erdos_cli = _build_erdos_client(config_obj)
                await erdos_cli.update_order_status(
                    id_pedido_erdos=order.external_order_id,
                    status="preparando",
                )
                await erdos_cli.close()
        except Exception:
            pass  # não bloqueia o fluxo principal

    return {
        "status": "ok" if not errors else "partial",
        "horus_pedido_remessa": cod_ped_remessa,
        "horus_pedido_venda": cod_ped_venda,
        "horus_cod_cli_final": cod_cli_final,
        "errors": errors,
    }


# ==============================================================================
# VERIFICAR STATUS NO ERDOS (polling individual)
# ==============================================================================

@router.post("/orders/{company_id}/{order_id}/check-erdos-status")
async def check_erdos_status(
    company_id: int,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Consulta GET /pedidos/{id} no Erdos e sincroniza o status local.

    Fluxo de acompanhamento após PATCH "preparando":
      - status Erdos = "preparando"  → mantém SENT_TO_HORUS localmente
      - status Erdos = "postado"     → registra tracking se disponível
      - status Erdos = "cancelado"   → marca CANCELLED localmente
    Matriz de cenários completa:

    Erdos \Local   | PENDING           | SENT_TO_HORUS         | DISPATCHED        | CANCELLED
    --------------|-------------------|-----------------------|-------------------|----------
    aguardando    | OK (sem ação)     | inconsistência (log)  | inconsistência    | log
    preparando    | atualiza log      | OK (sem ação)         | OK                | inconsistência
    postado       | log               | captura rastreio      | OK                | log
    cancelado     | cancela local ✅  | ALERTA BLOQUEANTE ⛔  | log crítico       | sem ação
    entregue      | inconsistência    | log                   | OK (log)          | log
    """
    _require_seller_or_master(current_user, company_id)

    order = db.query(DropshipOrder).filter(
        DropshipOrder.id == order_id,
        DropshipOrder.company_id == company_id,
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido não encontrado.")

    config = db.query(DropshipConfig).filter(DropshipConfig.id == order.config_id).first()
    client = _build_erdos_client(config)

    try:
        erdos_data = await client.get_order(order.external_order_id)
    except ErdosClientError as e:
        raise HTTPException(status_code=502, detail=f"Erro ao consultar Erdos: {str(e)}")
    finally:
        await client.close()

    erdos_status = (erdos_data.get("status") or "").lower().strip()
    previous_erdos_status = order.erdos_status or ""
    previous_local_status = order.status

    action = "none"
    alert = None
    changed = erdos_status != previous_erdos_status

    # Helper para registrar logs
    def _append_log(event: str, detail: str):
        current_logs = list(order.logs or [])
        current_logs.append({
            "at": datetime.utcnow().isoformat(),
            "event": event,
            "erdos_status": erdos_status,
            "local_status_before": previous_local_status,
            "local_status_after": order.status,
            "detail": detail,
        })
        order.logs = current_logs

    # Sempre atualiza o status Erdos conhecido e o timestamp da última consulta
    order.erdos_status = erdos_status
    order.erdos_checked_at = datetime.utcnow()

    # ─────────────────────────────────────────────────────────────────────────
    # CENÁRIO: CANCELADO no Erdos
    # ─────────────────────────────────────────────────────────────────────────
    if erdos_status == "cancelado":

        if order.status == "CANCELLED":
            # Já cancelado localmente — sem ação
            action = "already_cancelled"

        elif order.status == "DISPATCHED":
            # Conflito grave: despachamos mas Erdos cancelou
            # Não podemos desfazer — apenas registrar para análise manual
            action = "conflict_dispatched_but_cancelled"
            alert = (
                "⚠️ CONFLITO CRÍTICO: pedido foi despachado pelo Cronuz mas consta como "
                "CANCELADO no Erdos. Contate o suporte da Erdos para resolução manual."
            )
            if changed:
                _append_log("CONFLICT_DISPATCHED_CANCELLED", alert)

        elif order.status == "SENT_TO_HORUS":
            # Cancelado DEPOIS de já ter sido enviado ao Hórus
            # → Não cancela automaticamente: usuário DEVE cancelar no Hórus primeiro
            order.erdos_alert = True
            action = "alert_cancel_in_horus"
            alert = (
                "⛔ Pedido CANCELADO no Erdos após envio ao Hórus. "
                "Cancele os pedidos #remessa e #venda no Hórus ERP antes de prosseguir. "
                "O status local só será atualizado após confirmação manual."
            )
            if changed:
                _append_log("CANCELLED_AFTER_HORUS_SENT", alert)

        else:
            # PENDING ou qualquer outro — cancela localmente de forma segura
            order.status = "CANCELLED"
            order.erdos_alert = False
            action = "auto_cancelled"
            _append_log("AUTO_CANCELLED", "Cancelado no Erdos sem envio ao Hórus. Status local atualizado automaticamente.")

    # ─────────────────────────────────────────────────────────────────────────
    # CENÁRIO: POSTADO — capturar rastreio
    # ─────────────────────────────────────────────────────────────────────────
    elif erdos_status == "postado":
        tracking = (
            erdos_data.get("codigo_rastreio")
            or erdos_data.get("codigo_rastreamento")
            or (erdos_data.get("logistica") or {}).get("codigo_rastreio")
        )
        if tracking and not order.tracking_code:
            order.tracking_code = tracking
            action = "tracking_captured"
            _append_log("TRACKING_CAPTURED", f"Código de rastreio capturado: {tracking}")
        elif changed:
            action = "status_updated"
            _append_log("STATUS_UPDATED", f"Status Erdos atualizado para '{erdos_status}'.")
        else:
            action = "none"

    # ─────────────────────────────────────────────────────────────────────────
    # CENÁRIO: ENTREGUE
    # ─────────────────────────────────────────────────────────────────────────
    elif erdos_status == "entregue":
        if order.status == "DISPATCHED":
            action = "delivered_ok"
            if changed:
                _append_log("DELIVERED", "Pedido confirmado como entregue no Erdos.")
        else:
            action = "inconsistency_delivered_not_dispatched"
            if changed:
                _append_log("INCONSISTENCY", f"Erdos marcou como entregue mas status local é '{order.status}'.")

    # ─────────────────────────────────────────────────────────────────────────
    # CENÁRIO: PREPARANDO — verificar consistência
    # ─────────────────────────────────────────────────────────────────────────
    elif erdos_status == "preparando":
        if order.status == "SENT_TO_HORUS" or order.status == "DISPATCHED":
            action = "consistent"
            if changed:
                _append_log("STATUS_UPDATED", "Erdos confirmou status 'preparando' — consistente.")
        elif order.status == "PENDING":
            # Inconsistência: Erdos diz preparando mas não enviamos ao Hórus
            action = "inconsistency_preparando_but_pending"
            if changed:
                _append_log("INCONSISTENCY", "Erdos status 'preparando' mas pedido ainda está PENDING localmente.")
        else:
            action = "status_updated"
            if changed:
                _append_log("STATUS_UPDATED", f"Erdos status '{erdos_status}', local '{order.status}'.")

    # ─────────────────────────────────────────────────────────────────────────
    # CENÁRIO: AGUARDANDO — estado normal de fila
    # ─────────────────────────────────────────────────────────────────────────
    elif erdos_status == "aguardando":
        if order.status in ["SENT_TO_HORUS", "DISPATCHED"]:
            # Inconsistência: enviamos ao Hórus mas Erdos ainda vê como aguardando
            # Pode ser que o PATCH preparando não chegou
            action = "inconsistency_sent_but_aguardando"
            if changed:
                _append_log("INCONSISTENCY", f"Pedido enviado ao Hórus mas Erdos ainda está '{erdos_status}'. PATCH preparando pode ter falhado.")
        else:
            action = "queue_ok"

    # ─────────────────────────────────────────────────────────────────────────
    # STATUS DESCONHECIDO
    # ─────────────────────────────────────────────────────────────────────────
    else:
        action = "unknown_status"
        if changed:
            _append_log("UNKNOWN_STATUS", f"Status Erdos desconhecido: '{erdos_status}'.")

    db.commit()

    return {
        "erdos_status": erdos_status,
        "previous_erdos_status": previous_erdos_status,
        "local_status": order.status,
        "action": action,
        "alert": alert,
        "erdos_alert": order.erdos_alert,
        "changed": changed,
        "tracking_code": order.tracking_code,
        "erdos_checked_at": order.erdos_checked_at.isoformat() if order.erdos_checked_at else None,
    }


# ==============================================================================
# CONFIRMAR DESPACHO
# ==============================================================================

@router.post("/orders/{company_id}/{order_id}/confirm-dispatch")
async def confirm_dispatch(
    company_id: int,
    order_id: int,
    payload: Optional[ConfirmDispatchRequest] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Confirma o despacho do pedido de dropship com verificação estrita de segurança no Hórus ERP:
    1. Consulta Busca_NotaFiscal no Hórus para o pedido de remessa (CFOP 6.923).
    2. Valida se a nota possui STATUS == 'SAÍDA' e se CHAVE_ACESSO_NFE foi emitida.
    3. Armazena CHAVE_ACESSO_NFE, NRO_NOTA_FISCAL e DAT_EMISSAO_NF localmente.
    4. Notifica o Hub-Erdos via POST /pedidos/atualizar-status-despacho.
    5. Atualiza o status do pedido local para DISPATCHED.
    """
    _require_seller_or_master(current_user, company_id)

    order = db.query(DropshipOrder).filter(
        DropshipOrder.id == order_id,
        DropshipOrder.company_id == company_id,
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido dropship não encontrado.")

    if not order.horus_pedido_remessa:
        raise HTTPException(
            status_code=400,
            detail="Não é possível confirmar despacho: Pedido ainda não possui o número de remessa gerado no Hórus ERP (horus_pedido_remessa)."
        )

    # 1. Consultar Busca_NotaFiscal no Hórus ERP
    settings = db.query(CompanySettings).filter(CompanySettings.company_id == company_id).first()
    config = db.query(DropshipConfig).filter(DropshipConfig.id == order.config_id).first()

    horus_client = HorusLogisticsClient(db, company_id)
    try:
        cod_empresa = settings.horus_company if settings and settings.horus_company else "1"
        cod_filial = settings.horus_branch if settings and settings.horus_branch else "1"
        cod_metodo = (config.horus_cod_metodo if config and config.horus_cod_metodo else None) or getattr(settings, 'horus_cod_metodo', None) or 2

        # Trata COD_PED_VENDA numérico estrito (ex: 61 ou 650693)
        remessa_num = re.sub(r'\D', '', str(order.horus_pedido_remessa))
        if not remessa_num:
            raise HTTPException(status_code=400, detail="Código de remessa inválido para consulta de NF no Hórus.")

        nf_params = {
            "COD_EMPRESA": str(cod_empresa),
            "COD_FILIAL": str(cod_filial),
            "COD_PED_VENDA": int(remessa_num),
            "COD_METODO": cod_metodo,
            "OFFSET": 0,
            "LIMIT": 10
        }
        if order.horus_cod_cli_final:
            nf_params["COD_CLI"] = order.horus_cod_cli_final

        res_nf = await horus_client.get("Busca_NotaFiscal", params=nf_params)
        await horus_client.close()

        if not res_nf or (isinstance(res_nf, list) and len(res_nf) == 0):
            raise HTTPException(
                status_code=400,
                detail=f"Nenhuma Nota Fiscal localizada no Hórus ERP para o pedido de remessa #{remessa_num} (COD_METODO: {cod_metodo}). Fature a nota no Hórus antes de confirmar o despacho."
            )

        nf_list = res_nf if isinstance(res_nf, list) else [res_nf]

        # Procurar nota fiscal com STATUS == "SAÍDA" / "SAIDA"
        saida_nf = None
        for item in nf_list:
            if isinstance(item, dict):
                st = str(item.get("STATUS", "")).strip().upper()
                if "SA" in st and "DA" in st:  # SAÍDA / SAIDA
                    saida_nf = item
                    break

        if not saida_nf:
            # Fallback: primeira nota se possuir CHAVE_ACESSO_NFE
            first = nf_list[0]
            if isinstance(first, dict) and first.get("CHAVE_ACESSO_NFE"):
                saida_nf = first

        if not saida_nf or not saida_nf.get("CHAVE_ACESSO_NFE"):
            raise HTTPException(
                status_code=400,
                detail=f"A Nota Fiscal de Remessa #{remessa_num} ainda não foi faturada no Hórus com STATUS 'SAÍDA' ou não possui a CHAVE_ACESSO_NFE gerada."
            )

        chave_acesso = str(saida_nf.get("CHAVE_ACESSO_NFE", "")).strip()
        nro_nota = str(saida_nf.get("NRO_NOTA_FISCAL", "")).strip()
        dat_emissao = str(saida_nf.get("DAT_EMISSAO_NF", "")).strip()

        if not chave_acesso:
            raise HTTPException(
                status_code=400,
                detail="Segurança de Despacho: A Nota Fiscal localizada no Hórus ERP não possui CHAVE_ACESSO_NFE válida."
            )

    except HTTPException:
        raise
    except Exception as e:
        await horus_client.close()
        raise HTTPException(status_code=400, detail=f"Erro ao consultar Busca_NotaFiscal no Hórus ERP: {str(e)}")

    # 2. Notificar o Hub-Erdos via API (POST /pedidos/atualizar-status-despacho)
    erdos_client = _build_erdos_client(config)
    tracking_code = (payload.tracking_code if payload else None) or order.tracking_code or ""
    erdos_response = None

    try:
        erdos_response = await erdos_client.confirm_dispatch(
            id_pedido_erdos=order.external_order_id,
            tracking_code=tracking_code,
            chave_nfe_remessa=chave_acesso,
        )
    except ErdosClientError as e:
        raise HTTPException(
            status_code=502,
            detail=f"Nota Fiscal capturada no Hórus (NF {nro_nota}), porém ocorreu erro ao comunicar despacho com o Hub-Erdos: {str(e)}"
        )
    finally:
        await erdos_client.close()

    # 3. Atualizar dados do pedido localmente com total segurança
    status_before = order.status
    order.status = "DISPATCHED"
    order.nfe_remessa_key = chave_acesso
    if tracking_code:
        order.tracking_code = tracking_code
    order.dispatched_at = datetime.utcnow()

    # Atualizar objeto JSON fiscal_data
    fdata = dict(order.fiscal_data or {})
    fdata["nro_nota_fiscal_remessa"] = nro_nota
    fdata["dat_emissao_nf_remessa"] = dat_emissao
    fdata["chave_nfe_remessa_6923"] = chave_acesso
    order.fiscal_data = fdata

    # Registrar histórico de log
    logs = list(order.logs or [])
    logs.append({
        "at": datetime.utcnow().isoformat(),
        "event": "CONFIRM_DISPATCH",
        "local_status_before": status_before,
        "local_status_after": "DISPATCHED",
        "detail": f"NF-e de Remessa nº {nro_nota} (Chave: {chave_acesso}) capturada com sucesso do Hórus. Despacho verificado e confirmado no Erdos."
    })
    order.logs = logs

    db.commit()
    db.refresh(order)

    return {
        "status": "dispatched",
        "nro_nota_fiscal": nro_nota,
        "chave_nfe_remessa_6923": chave_acesso,
        "dat_emissao_nf": dat_emissao,
        "erdos_response": erdos_response
    }


# ==============================================================================
# DOWNLOAD DE DOCUMENTOS
# ==============================================================================

@router.get("/orders/{company_id}/{order_id}/documents/{doc_type}")
async def get_order_document(
    company_id: int,
    order_id: int,
    doc_type: str,  # xml | danfe | etiqueta
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Retorna URL pública do documento armazenado localmente.
    doc_type: xml | danfe | etiqueta
    """
    _require_seller_or_master(current_user, company_id)

    order = db.query(DropshipOrder).filter(
        DropshipOrder.id == order_id,
        DropshipOrder.company_id == company_id,
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Pedido não encontrado.")

    path_map = {
        "xml": order.xml_path,
        "danfe": order.danfe_path,
        "etiqueta": order.label_path,
    }

    if doc_type not in path_map:
        raise HTTPException(status_code=400, detail="Tipo de documento inválido. Use: xml, danfe ou etiqueta.")

    local_path = path_map[doc_type]
    if not local_path:
        raise HTTPException(status_code=404, detail=f"Documento '{doc_type}' não disponível para este pedido.")

    # Normalizar: o path pode estar salvo como relativo (ex: "uploads/...") ou absoluto
    # Resolver sempre relativo ao diretório raiz do backend
    import pathlib
    backend_dir = pathlib.Path(__file__).resolve().parent.parent.parent  # app/api -> app -> backend root
    abs_path = pathlib.Path(local_path)
    if not abs_path.is_absolute():
        abs_path = backend_dir / local_path

    if not abs_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Arquivo '{doc_type}' não encontrado em disco. Tente sincronizar o pedido novamente."
        )

    # Retornar URL pública via /uploads montado no main.py
    # Normalizar para sempre começar com "uploads/..."
    local_str = str(abs_path)
    uploads_marker = "uploads/"
    idx = local_str.find(uploads_marker)
    if idx >= 0:
        public_url = "/" + local_str[idx:]
    else:
        public_url = f"/uploads/{company_id}/dropship/{order.external_order_id}/{abs_path.name}"

    return {"url": public_url, "path": str(abs_path)}


# ==============================================================================
# ESTOQUE
# ==============================================================================

@router.post("/stock/{company_id}/push")
async def push_stock_to_hub(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Consulta estoque no Hórus via AcervoB2B (filtrando por COD_BARRA_ITEM/ISBN)
    e envia posição atualizada para o Hub-Erdos.
    Campo de vínculo: COD_BARRA_ITEM (ISBN-13) ↔ sku do Hub-Erdos.
    """
    _require_seller_or_master(current_user, company_id)

    config = _get_config_or_404(db, company_id)
    client = _build_erdos_client(config)

    if not config.horus_customer:
        await client.close()
        raise HTTPException(status_code=400, detail="Customer parceiro não configurado.")

    try:
        from app.integrators.horus_products import HorusProducts
        horus_prod = HorusProducts(db, company_id)

        customer = config.horus_customer
        # Busca acervo completo do seller no Hórus (paginação ampla)
        result = await horus_prod.busca_acervo_b2b(
            id_doc=customer.id_doc,
            id_guid=customer.id_guid,
            limit=10000,
            offset=0,
        )

        items_to_send = []
        if isinstance(result, list):
            for item in result:
                isbn = item.get("COD_BARRA_ITEM") or item.get("BARRAS_ISBN") or item.get("ISBN")
                saldo = item.get("SALDO") or item.get("QTD_SALDO") or item.get("QTD_DISPONIVEL") or 0
                if isbn and str(isbn).strip():
                    items_to_send.append({
                        "sku": str(isbn).strip(),
                        "quantidade": max(0, int(saldo))
                    })

        if not items_to_send:
            await horus_prod.close()
            await client.close()
            return {"status": "warning", "message": "Nenhum item encontrado no Hórus para enviar.", "skus": 0}

        push_result = await client.push_stock(items_to_send)

        # Atualizar last_run
        config.stock_sync_last_run = datetime.utcnow()
        db.commit()

        await horus_prod.close()
        await client.close()

        return {
            "status": "ok",
            "skus_sent": len(items_to_send),
            "hub_response": push_result,
        }

    except ErdosClientError as e:
        raise HTTPException(status_code=502, detail=f"Erro ao enviar estoque para Hub-Erdos: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro interno ao processar estoque: {str(e)}")


@router.get("/stock/{company_id}/hub")
async def get_hub_stock(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Consulta posição de estoque registrada no Hub-Erdos para este fornecedor."""
    _require_seller_or_master(current_user, company_id)

    config = _get_config_or_404(db, company_id)
    client = _build_erdos_client(config)
    try:
        result = await client.get_stock()
        return result
    except ErdosClientError as e:
        raise HTTPException(status_code=502, detail=str(e))
    finally:
        await client.close()


# ==============================================================================
# TABELA DE PREÇOS DROPSHIP (dsp_price_table)
# ==============================================================================

class PriceTableItemResponse(BaseModel):
    id: int
    company_id: int
    isbn: str
    titulo: Optional[str]
    desconto: float
    data_validade: date
    created_at: Optional[datetime]
    updated_at: Optional[datetime]
    vencido: bool = False

    class Config:
        from_attributes = True


@router.get("/price-table/{company_id}", response_model=List[PriceTableItemResponse])
async def list_price_table(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista todos os itens da tabela de preços do seller."""
    _require_seller_or_master(current_user, company_id)
    items = (
        db.query(DropshipPriceTable)
        .filter(DropshipPriceTable.company_id == company_id)
        .order_by(DropshipPriceTable.data_validade.desc(), DropshipPriceTable.isbn)
        .all()
    )
    today = date.today()
    result = []
    for item in items:
        result.append(PriceTableItemResponse(
            id=item.id,
            company_id=item.company_id,
            isbn=item.isbn,
            titulo=item.titulo,
            desconto=float(item.desconto),
            data_validade=item.data_validade,
            created_at=item.created_at,
            updated_at=item.updated_at,
            vencido=item.data_validade < today,
        ))
    return result


@router.post("/price-table/{company_id}/upload")
async def upload_price_table(
    company_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Faz upload de planilha (.xlsx ou .csv) com colunas:
    isbn, titulo, desconto, data_validade.
    Realiza upsert por (company_id, isbn) — sobrescreve se já existir.
    """
    _require_seller_or_master(current_user, company_id)

    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Formato inválido. Envie um arquivo .xlsx ou .csv.")

    content = await file.read()

    rows: list = []
    try:
        if filename.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active

            # ── Leitura robusta de headers ──────────────────────────────────────
            # Aceita headers em português/abreviados e colunas sem nome (por posição).
            # Planilha esperada (com ou sem header nas primeiras colunas):
            #   Col A: isbn   | Col B: titulo | Col C: desconto | Col D: data_validade
            KNOWN_ALIASES = {
                # isbn
                "isbn": "isbn", "ean": "isbn", "codigo": "isbn", "código": "isbn",
                # titulo
                "titulo": "titulo", "título": "titulo", "nome": "titulo", "descricao": "titulo",
                "descrição": "titulo", "livro": "titulo",
                # desconto
                "desconto": "desconto", "desc": "desconto", "discount": "desconto",
                "perc": "desconto", "%": "desconto",
                # data_validade
                "data_validade": "data_validade", "validade": "data_validade",
                "vigencia": "data_validade", "vigência": "data_validade",
                "data": "data_validade", "expira": "data_validade", "vencimento": "data_validade",
                "data vigencia": "data_validade", "data vigência": "data_validade",
            }
            POSITIONAL_MAP = {0: "isbn", 1: "titulo", 2: "desconto", 3: "data_validade"}

            raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            # Resolve headers: por alias primeiro, por posição se vazio/None
            headers: list = []
            for idx_h, hval in enumerate(raw_headers):
                h_clean = str(hval or "").strip().lower()
                # Remove acentos simples para comparação
                h_norm = h_clean.replace("ê", "e").replace("ã", "a").replace("ç", "c").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
                resolved = KNOWN_ALIASES.get(h_clean) or KNOWN_ALIASES.get(h_norm)
                if not resolved and idx_h in POSITIONAL_MAP:
                    resolved = POSITIONAL_MAP[idx_h]
                headers.append(resolved or h_clean or f"col_{idx_h}")

            log.info(f"[PriceTable] Headers resolvidos: {headers}")

            for row in ws.iter_rows(min_row=2, values_only=True):
                # Ignora linhas completamente vazias
                if all(v is None for v in row):
                    continue
                rows.append(dict(zip(headers, row)))
        else:  # csv
            import csv
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append({k.strip().lower(): v for k, v in row.items()})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler o arquivo: {e}")


    imported = 0
    errors: list = []

    for idx, row in enumerate(rows, start=2):  # linha 2 = primeira linha de dados
        try:
            isbn_val = row.get("isbn")
            # ISBN pode vir como int ou float do openpyxl (ex: 9786559671977 ou 9786559671977.0)
            if isinstance(isbn_val, float):
                isbn_raw = str(int(isbn_val))
            elif isinstance(isbn_val, int):
                isbn_raw = str(isbn_val)
            else:
                isbn_raw = str(isbn_val or "").strip()

            # Ignora linhas vazias ou com ISBN inválido
            if not isbn_raw or isbn_raw.lower() in ("none", "", "nan"):
                errors.append({"linha": idx, "erro": "ISBN vazio — linha ignorada"})
                continue

            titulo_raw = str(row.get("titulo") or "").strip() or None
            desconto_raw = row.get("desconto") or row.get("desc") or 0
            validade_raw = row.get("data_validade") or row.get("validade") or row.get("data") or ""

            # Parse desconto
            try:
                desconto_val = float(str(desconto_raw).replace(",", ".").strip())
            except Exception:
                errors.append({"linha": idx, "isbn": isbn_raw, "erro": f"Desconto inválido: {desconto_raw!r}"})
                continue

            # Parse data_validade
            validade_parsed: Optional[date] = None
            if validade_raw:
                if hasattr(validade_raw, "date"):  # já é datetime do openpyxl
                    validade_parsed = validade_raw.date() if hasattr(validade_raw, "date") else validade_raw
                else:
                    validade_str = str(validade_raw).strip()
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                        try:
                            from datetime import datetime as dt_
                            validade_parsed = dt_.strptime(validade_str, fmt).date()
                            break
                        except ValueError:
                            continue

            if not validade_parsed:
                errors.append({"linha": idx, "isbn": isbn_raw, "erro": f"Data de validade inválida: {validade_raw!r}"})
                continue

            # Upsert por (company_id, isbn)
            existing = db.query(DropshipPriceTable).filter(
                DropshipPriceTable.company_id == company_id,
                DropshipPriceTable.isbn == isbn_raw,
            ).first()

            if existing:
                existing.titulo = titulo_raw
                existing.desconto = desconto_val
                existing.data_validade = validade_parsed
            else:
                db.add(DropshipPriceTable(
                    company_id=company_id,
                    isbn=isbn_raw,
                    titulo=titulo_raw,
                    desconto=desconto_val,
                    data_validade=validade_parsed,
                ))
            imported += 1
        except Exception as e:
            errors.append({"linha": idx, "erro": str(e)})

    db.commit()

    return {
        "importados": imported,
        "erros": len(errors),
        "detalhes_erros": errors,
    }


@router.delete("/price-table/{company_id}/{item_id}")
async def delete_price_table_item(
    company_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Remove um item específico da tabela de preços."""
    _require_seller_or_master(current_user, company_id)
    item = db.query(DropshipPriceTable).filter(
        DropshipPriceTable.id == item_id,
        DropshipPriceTable.company_id == company_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado.")
    db.delete(item)
    db.commit()
    return {"ok": True}


@router.delete("/price-table/{company_id}/clear/all")
async def clear_price_table(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Remove todos os itens da tabela de preços do seller."""
    _require_seller_or_master(current_user, company_id)
    deleted = db.query(DropshipPriceTable).filter(
        DropshipPriceTable.company_id == company_id
    ).delete()
    db.commit()
    return {"removidos": deleted}
